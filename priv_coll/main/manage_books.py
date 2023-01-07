from collections import namedtuple
from .api_request import get_api_request
from .models import Book
from openpyxl import load_workbook


def import_books_from_file(user: str, file: str = "main/source/source.xlsx") -> str:
    """add books stored in xls file"""

    count_success = 0
    wb = load_workbook(file)
    ws = wb.active
    row = "2"
    while ws["A" + row].value:
        print(ws["A" + row].value)
        b = Book()
        b.user = user
        b.ISBN = ws["A" + row].value
        b.title = ws["B" + row].value
        b.authors = ws["C" + row].value
        b.publishedDate = ws["D" + row].value
        b.pageCount = ws["E" + row].value if isinstance(ws["E" + row].value, int) else None
        b.thumbnail = ws["F" + row].value
        b.language = ws["G" + row].value
        b.cost = ws["H" + row].value
        if not Book.objects.filter(ISBN=b.ISBN, user=b.user):
            b.save()
            count_success += 1
        row = str(int(row) + 1)
    return f"successfully added {count_success} books"


def add_books(book_params: dict, user: str) -> namedtuple:
    """add books based on response from api"""

    response_dict = get_api_request(book_params)
    if response_dict == {"error": -1}:
        return namedtuple("Result", ["errors", "duplicates", "success"])(-1, -1, -1)

    count_errors = 0
    count_duplicates = 0
    count_success = 0

    for info in response_dict["items"]:
        vol = info.get("volumeInfo")
        curr_book = Book()
        curr_book.user = user

        # ISBN_13
        if not vol.get("industryIdentifiers"):  # No ISBN identification
            count_errors += 1
            continue
        for isbn_elem in vol.get("industryIdentifiers"):
            if isbn_elem.get("type") == "ISBN_13":
                curr_book.ISBN = isbn_elem.get("identifier")
        if curr_book.ISBN is None:  # No ISBN code
            count_errors += 1
            continue

        if Book.objects.filter(ISBN=curr_book.ISBN, user=user):  # ISBN already exists
            count_duplicates += 1
            continue

        curr_book.title = vol.get("title")

        if vol.get("authors"):
            if vol.get("authors") == 1:
                curr_book.authors = vol.get("authors")[0]
            else:
                curr_book.authors = " ".join(vol.get("authors"))

        curr_book.publishedDate = vol.get("publishedDate", None)
        curr_book.pageCount = vol.get("pageCount", None)
        curr_book.thumbnail = vol.get("imageLinks", {}).get("thumbnail", None)
        curr_book.language = vol.get("language", None)

        if info.get("saleInfo").get("saleability") == "FOR_SALE":
            curr_book.cost = info.get("saleInfo").get("listPrice").get("amount")
        else:
            curr_book.cost = None

        curr_book.save()

    return namedtuple("Result", ["errors", "duplicates", "success"])(count_errors, count_duplicates, count_success)


def edit_book(book_target_isbn: int, user: str, updated_book: dict) -> str:
    """execute changes for modified book"""

    b = Book.objects.get(ISBN=book_target_isbn, user=user)
    b.title = updated_book["title"]
    b.authors = updated_book["authors"]
    b.publishedDate = updated_book["publishedDate"]
    b.pageCount = updated_book["pageCount"]
    b.thumbnail = updated_book["thumbnail"]
    b.language = updated_book["language"]
    b.cost = updated_book["cost"]
    b.save()

    return "succesfully edited book"
